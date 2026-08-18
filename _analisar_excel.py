"""Script temporário para analisar a estrutura do Excel."""
import openpyxl

wb = openpyxl.load_workbook('Matriz_Malha_Logistica.xlsx', read_only=True, data_only=True)
print('Abas:', wb.sheetnames)
print()

ws = wb['Consolidado']
print(f'Consolidado: {ws.max_row} linhas x {ws.max_column} colunas')
print()

# Cabecalho
print('CABECALHO (linha 1):')
for col in range(1, min(ws.max_column + 1, 30)):
    val = ws.cell(1, col).value
    if val:
        letter = chr(64+col) if col <= 26 else str(col)
        print(f'  Col {col} ({letter}): {val}')
print()

# Primeiras linhas - colunas C e I
print('Amostra (Col C = Deposito, Col I = CEP):')
depositos_vistos = set()
for row in range(2, min(ws.max_row + 1, 200)):
    dep = ws.cell(row, 3).value  # C
    cep = ws.cell(row, 9).value  # I
    if dep and dep not in depositos_vistos:
        depositos_vistos.add(dep)
        print(f'  Linha {row}: Deposito="{dep}", CEP="{cep}"')

print(f'\nDepositos unicos (primeiras 200 linhas): {len(depositos_vistos)}')

# Todos os depositos unicos
all_deps = {}
for row in range(2, ws.max_row + 1):
    dep = ws.cell(row, 3).value
    cep = ws.cell(row, 9).value
    if dep and dep not in all_deps:
        all_deps[dep] = cep

print(f'Total de depositos unicos (todas as linhas): {len(all_deps)}')
print('\nLista completa de depositos com CEPs:')
for d in sorted(all_deps.keys()):
    print(f'  - "{d}" -> CEP: "{all_deps[d]}"')

wb.close()
