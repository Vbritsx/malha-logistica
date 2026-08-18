import openpyxl
import os

def analisar():
    file_path = "Matriz_Malha_Logistica.xlsx"
    if not os.path.exists(file_path):
        print(f"Erro: {file_path} não encontrado.")
        return
        
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb["Consolidado"]
    
    # Obter cabeçalho (linha 1)
    headers = []
    for col in range(1, ws.max_column + 1):
        headers.append(ws.cell(1, col).value)
        
    print("\n--- CABEÇALHO ATUALIZADO ---")
    for i, h in enumerate(headers):
        print(f"Coluna {i+1} ({chr(65+i) if i < 26 else i+1}): {h}")
        
    print("\n--- AMOSTRA DA LINHA 2 ---")
    row_values = [ws.cell(2, col).value for col in range(1, len(headers) + 1)]
    print(dict(zip(headers, row_values)))
    
    # Analisar CDs e seus endereços
    idx_cd_code = headers.index("Deposito")
    idx_cd_name = headers.index("Nome deposito")
    idx_cd_addr = headers.index("Endereço Deposito")
    idx_cd_city = headers.index("Cidade deposito")
    
    cds = {}
    for r in range(2, 50): # Amostra
        cd_code = ws.cell(r, idx_cd_code + 1).value
        cd_name = ws.cell(r, idx_cd_name + 1).value
        cd_addr = ws.cell(r, idx_cd_addr + 1).value
        cd_city = ws.cell(r, idx_cd_city + 1).value
        
        if cd_code and cd_code not in cds:
            cds[cd_code] = {
                "name": cd_name,
                "address": cd_addr,
                "city": cd_city
            }
            
    print("\n--- AMOSTRA DE CDS E ENDEREÇOS ---")
    for code, info in cds.items():
        print(f"CD: {code} | Nome: {info['name']} | Endereço: {info['address']} | Cidade: {info['city']}")
        
    wb.close()

if __name__ == "__main__":
    analisar()
