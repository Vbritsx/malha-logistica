import openpyxl
import os

def pesquisar_cds():
    file_path = "Matriz_Malha_Logistica.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    print("--- BUSCA DE REFERÊNCIAS DOS CDS ---")
    
    # 1. Verificar CD D002 na aba Consolidado para ver se há algum nome diferente
    ws = wb["Consolidado"]
    for row in range(2, ws.max_row + 1):
        dep = ws.cell(row, 4).value # Deposito (D002)
        if dep == "D002":
            label = ws.cell(row, 2).value
            nome = ws.cell(row, 3).value
            cidade = ws.cell(row, 5).value
            print(f"D002 em Consolidado - Label: {label} | Nome: {nome} | Cidade: {cidade}")
            break
            
    # 2. Procurar menções de "D002", "D019", "ABV", "Lapa" nas outras abas
    for name in ["Fuzzy", "FuzzyLookup_AddIn_Undo_Sheet", "Clusters", "rascunhos"]:
        ws_sheet = wb[name]
        print(f"\nBuscando na aba: {name}...")
        encontrados = 0
        for r in range(1, min(ws_sheet.max_row + 1, 1000)):  # Limitar busca para não demorar
            row_vals = [str(ws_sheet.cell(r, c).value) for c in range(1, ws_sheet.max_column + 1)]
            row_str = " | ".join(row_vals)
            if any(x in row_str for x in ["D002", "D019", "D024", "ABV", "LAPA"]):
                print(f"  Linha {r}: {row_str[:150]}")
                encontrados += 1
                if encontrados >= 5:
                    print("  ... mais resultados omitidos.")
                    break
                    
    wb.close()

if __name__ == "__main__":
    pesquisar_cds()
